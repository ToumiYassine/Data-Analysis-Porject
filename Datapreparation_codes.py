from openpyxl import load_workbook, Workbook

# Charger le fichier Excel existant
wb = load_workbook('test/test1.xlsx')
ws = wb.active

# Créer un nouveau fichier Excel pour sauvegarder le résultat
new_wb = Workbook()
new_ws = new_wb.active

# Copier l'entête (si présent)
header = [cell.value for cell in ws[1]]
new_ws.append(header)

# Parcourir les lignes du fichier source à partir de la deuxième ligne
for row in ws.iter_rows(min_row=2, values_only=True):
    # Vérifier si la valeur de la colonne B commence par '1005' ou '1001'
    if str(row[0])[:4] in ('1005', '1001'):
        new_ws.append(row)

# Sauvegarder le nouveau fichier Excel
new_wb.save('test/test2.xlsx')

print("Le fichier a été sauvegardé avec succès sous 'test2.xlsx'.")
