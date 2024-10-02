from openpyxl import load_workbook
from datetime import datetime

# Dictionnaire pour convertir les mois en espagnol en chiffres
mois_espagnol = {
    'Enero': '01',
    'Febrero': '02',
    'Marzo': '03',
    'Abril': '04',
    'Mayo': '05',
    'Junio': '06',
    'Julio': '07',
    'Agosto': '08',
    'Septiembre': '09',
    'Octubre': '10',
    'Noviembre': '11',
    'Diciembre': '12'
}

# Charger le fichier Excel existant
wb = load_workbook('test/test2.xlsx')
ws = wb.active

# Parcourir les colonnes de B à M pour mettre à jour les en-têtes
for col in ws.iter_cols(min_col=2, max_col=13, min_row=1, max_row=1):
    for cell in col:
        # Extraire le mois et l'année de l'en-tête
        if '-' in cell.value:
            mois, annee = cell.value.split(' - ')
            mois = mois.strip()  # Enlever les espaces autour du mois

            # Convertir le mois en chiffre en utilisant le dictionnaire
            if mois in mois_espagnol:
                mois_numero = mois_espagnol[mois]
                # Créer une date avec le format requis
                nouvelle_date = f"01/{mois_numero}/{annee} 00:00:00"
                # Convertir en objet datetime pour validation
                date_formattee = datetime.strptime(nouvelle_date, "%d/%m/%Y %H:%M:%S")
                # Mettre à jour la cellule avec la date formatée
                cell.value = date_formattee

# Sauvegarder les modifications dans un nouveau fichier
wb.save('test/test3.xlsx')

print("Les en-têtes ont été modifiées et le fichier a été sauvegardé sous 'test3.xlsx'.")
