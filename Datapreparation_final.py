import openpyxl

# Charger le fichier Excel
wb = openpyxl.load_workbook('test/test3.xlsx')
sheet = wb.active

# Liste pour stocker les nouvelles lignes
new_data = []

# Boucle sur chaque ligne de pays (ici à partir de la ligne 2 jusqu'à la dernière ligne de données)
for row in range(2, sheet.max_row + 1):
    # Extraire le pays et le nom du book
    pais = sheet.cell(row=row, column=1).value
    book = sheet.cell(row=row, column=sheet.max_column).value  # Dernière colonne pour 'Country'

    # Parcourir les colonnes contenant les dates (sauf la dernière qui est pour 'Country')
    for col in range(2, sheet.max_column):
        # Extraire la date (en première ligne de chaque colonne)
        date = sheet.cell(row=1, column=col).value
        # Extraire la valeur pour ce pays à cette date
        valeur = sheet.cell(row=row, column=col).value
        # Ajouter la nouvelle ligne dans la liste
        new_data.append([pais, valeur, date, book])

# Créer une nouvelle feuille ou un nouveau fichier Excel
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.title = "Transformed Data"

# Ajouter l'entête
new_sheet.append(["Code", "Value", "Date", "Country"])

# Ajouter les nouvelles données transformées
for row in new_data:
    new_sheet.append(row)

# Sauvegarder le nouveau fichier Excel
new_wb.save("test/test4.xlsx")
