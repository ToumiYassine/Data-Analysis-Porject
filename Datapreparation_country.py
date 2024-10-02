from openpyxl import load_workbook, Workbook

# Charger le fichier Excel existant
url="Data_scraped/bd2024.xlsx"
wb = load_workbook(url)
ws = wb.active

# Créer un nouveau fichier Excel pour sauvegarder le résultat
new_wb = Workbook()
new_ws = new_wb.active

# Ajouter un en-tête pour la nouvelle colonne "Pays"
header = [cell.value for cell in ws[1]] + ['Pays']
new_ws.append(header)

# Variable pour garder le nom du pays courant
current_country = ""

# Parcourir les lignes du fichier source à partir de la deuxième ligne
for row in ws.iter_rows(min_row=2, values_only=True):
    # Vérifier si la cellule de la colonne A contient un nom de pays (pas un chiffre)
    if isinstance(row[0], str) and not row[0].isdigit():
        current_country = row[0]  # Mettre à jour le nom du pays courant
    else:
        # Ajouter le pays correspondant à chaque ligne de données
        new_row = list(row) + [current_country]
        new_ws.append(new_row)
new_wb.save('test/test1.xlsx')
print("Le fichier a été sauvegardé avec succès sous 'test1.xlsx'.")
